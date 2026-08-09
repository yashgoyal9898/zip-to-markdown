# zip_to_md.py

import csv
from io import BytesIO, StringIO
from pathlib import Path
import zipfile

import openpyxl


def extract_zip_to_markdown(zip_bytes: BytesIO) -> str:

    with zipfile.ZipFile(zip_bytes, 'r') as z:
        files = [f for f in z.namelist() if not f.endswith('/')]

        md_content = []
        md_content.append("# ZIP Content Dump\n\n")

        for file in files:
            md_content.append("\n---\n")
            md_content.append(f"## 📄 {file}\n")

            try:
                content = z.read(file)
                ext = Path(file).suffix.lower()

                if ext == ".csv":
                    md_content.append(_csv_to_markdown(content))
                elif ext == ".xlsx":
                    md_content.append(_xlsx_to_markdown(content))
                else:
                    text = _safe_decode(content)
                    lang = _detect_language(file)
                    md_content.append(f"```{lang}\n{text}\n```\n")

            except Exception as e:
                md_content.append(f"`Error reading file: {e}`\n")

        return "".join(md_content)


# ---------- INTERNAL HELPERS ---------- #

def _safe_decode(content: bytes) -> str:
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("latin-1")


def _detect_language(file_path: str) -> str:
    ext = Path(file_path).suffix.lstrip(".")
    return ext if ext else ""


def _escape_pipe(text: str) -> str:
    """Escape pipe and newline characters so markdown table cells don't break."""
    return text.replace("|", "\\|").replace("\n", " ").replace("\r", "")


def _csv_to_markdown(content: bytes) -> str:
    text = _safe_decode(content)
    reader = csv.reader(StringIO(text))
    rows = list(reader)

    if not rows:
        return "_Empty CSV_\n"

    header = [_escape_pipe(cell) for cell in rows[0]]
    col_count = len(header)

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * col_count) + " |",
    ]
    for row in rows[1:]:
        cells = [_escape_pipe(cell) for cell in row]
        # Pad short rows to match header width
        while len(cells) < col_count:
            cells.append("")
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines) + "\n"


def _xlsx_to_markdown(content: bytes) -> str:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"### Sheet: {sheet_name}\n")

        rows = list(ws.iter_rows(values_only=True))

        # Strip trailing empty rows
        while rows and all(cell is None for cell in rows[-1]):
            rows.pop()

        if not rows:
            parts.append("_Empty sheet_\n")
            continue

        header = [_escape_pipe(str(cell) if cell is not None else "") for cell in rows[0]]
        col_count = len(header)

        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * col_count) + " |",
        ]
        for row in rows[1:]:
            cells = [_escape_pipe(str(cell) if cell is not None else "") for cell in row]
            lines.append("| " + " | ".join(cells) + " |")

        parts.append("\n".join(lines) + "\n")

    return "\n".join(parts)
